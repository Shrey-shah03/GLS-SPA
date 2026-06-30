from flask import Flask, request, jsonify, send_from_directory, render_template, Response
import os
import re
import json
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import matching_engine

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Support for Render mounted disks or other persistent volumes
DATA_DIR = os.environ.get("DATA_DIR", BASE_DIR)

app = Flask(__name__, static_folder="static", template_folder="templates")

@app.before_request
def require_auth():
    admin_password = os.environ.get("ADMIN_PASSWORD")
    if admin_password:
        auth = request.authorization
        if not auth or auth.username != 'admin' or auth.password != admin_password:
            return Response('Login Required', 401, {'WWW-Authenticate': 'Basic realm="Login Required"'})

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
GENERATED_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generated")
CATALOG_JSON_PATH = os.path.join(DATA_DIR, "catalog_text.json")
BUYERS_DB_PATH = os.path.join(DATA_DIR, "buyers_db.json")

# Initialize persistent data on first boot if using a mounted volume
def init_persistent_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    for filename in ["catalog_text.json", "buyers_db.json", "GLS_SPA_Product_Database.csv", "catalog_specs.json"]:
        src = os.path.join(BASE_DIR, filename)
        dst = os.path.join(DATA_DIR, filename)
        if os.path.exists(src) and not os.path.exists(dst):
            shutil.copy2(src, dst)

init_persistent_data()

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(GENERATED_FOLDER, exist_ok=True)

def detect_columns(row_vals):
    desc_col_idx = None
    qty_col_idx = None
    
    exclude_desc = [
        "hsn", "mat code", "material code", "color", "angle", "qty", "quantity", 
        "unit", "uom", "price", "rate", "amount", "cost", "make", "brand", 
        "manufacturer", "driver", "led", "image", "picture", "photo", 
        "location", "area", "sr.", "s.no", "serial", "remark", "status", 
        "inspection", "packing", "stickering", "branding", "watt", "temp", 
        "cct", "dimension", "size", "height", "cutout", "dia", "voltage"
    ]
    
    desc_priority_1 = ["description", "item name", "item description", "particulars", "specification"]
    desc_priority_2 = ["item", "luminaire", "product description", "particular"]
    desc_priority_3 = ["product", "name", "model", "code", "type"]
    
    # First, find Qty column
    for c_idx, val in enumerate(row_vals):
        if val:
            val_lower = str(val).lower().strip()
            if val_lower == "qty" or val_lower == "quantity":
                qty_col_idx = c_idx + 1
                break
                
    if qty_col_idx is None:
        for c_idx, val in enumerate(row_vals):
            if val:
                val_lower = str(val).lower().strip()
                if "qty" in val_lower or "quantity" in val_lower:
                    qty_col_idx = c_idx + 1
                    break

    # Now, find Description column
    for priority in [desc_priority_1, desc_priority_2, desc_priority_3]:
        for c_idx, val in enumerate(row_vals):
            if val:
                val_lower = str(val).lower().strip()
                is_excluded = any(ex in val_lower for ex in exclude_desc)
                if not is_excluded:
                    if any(kw in val_lower for kw in priority):
                        desc_col_idx = c_idx + 1
                        return desc_col_idx, qty_col_idx
                        
    # Fallback to first non-excluded column
    for c_idx, val in enumerate(row_vals):
        if val:
            val_lower = str(val).lower().strip()
            is_excluded = any(ex in val_lower for ex in exclude_desc)
            if not is_excluded and c_idx + 1 != qty_col_idx:
                desc_col_idx = c_idx + 1
                break
                
    return desc_col_idx, qty_col_idx

def clean_gls_code(desc_str, specs_gls_code):
    desc_clean = str(desc_str).strip()
    
    # Find start of GS- or GS code
    match = re.search(r'\b(GS[- ].*)', desc_clean, re.IGNORECASE)
    if not match:
        # If no GS code, let's see if we can find any other code or just return first part
        parts = re.split(r'\s*-\s*|\s*,\s*', desc_clean)
        code = parts[0].strip()
    else:
        gs_part = match.group(1).strip()
        # Split by common separators and specifications keywords to isolate the code part
        parts = re.split(r'\s+-\s+|\s*,\s*|\b(?:wattage|size|material|length|dia|aluminium|special|beam|color|temp|uom|qty|hsn)\b', gs_part, flags=re.IGNORECASE)
        code = parts[0].strip()
        
    # Clean up any trailing hyphens, slashes, or spaces
    code = re.sub(r'[\-\s/]+$', '', code)
    
    # If specs_gls_code is a clean, standard code and is a substring of code, prefer specs_gls_code
    if specs_gls_code:
        specs_upper = str(specs_gls_code).strip().upper()
        code_upper = code.upper()
        if specs_upper in code_upper:
            return specs_gls_code
            
    if len(code) > 2:
        return code
        
    return specs_gls_code or desc_clean

def generate_description_from_boq(desc_str):
    desc_clean = str(desc_str).strip()
    desc_lower = desc_clean.lower()
    
    # 1. Handle Accessories / non-luminaires first (wires, caps, adapters, joints, drivers)
    # Check for End Cap
    if "end cap" in desc_lower or "endcap" in desc_lower or "ean cap" in desc_lower:
        if "track" in desc_lower or "channel" in desc_lower or "gs-mt" in desc_lower:
            return "End Cap for Magnetic Track Channel."
        return "End Cap."
        
    # Check for Live End / Power Adapter / Power Feed
    if "live end" in desc_lower or "power adapter" in desc_lower or "power adaptor" in desc_lower or "liveend" in desc_lower:
        if "track" in desc_lower or "channel" in desc_lower or "gs-mt" in desc_lower:
            return "Power Feed / Live End for Magnetic Track."
        return "Power Feed Adapter."
        
    # Check for Jointer / Joint
    if "jointer" in desc_lower or "joint" in desc_lower or "connector" in desc_lower:
        # Check type (I, L, T, corner, inline)
        type_str = ""
        if "i jointer" in desc_lower or "i-jointer" in desc_lower or "i joint" in desc_lower or "inline" in desc_lower:
            type_str = "I-Jointer "
        elif "l jointer" in desc_lower or "l-jointer" in desc_lower or "l joint" in desc_lower or "corner" in desc_lower:
            type_str = "L-Jointer "
        elif "t jointer" in desc_lower or "t-jointer" in desc_lower or "t joint" in desc_lower:
            type_str = "T-Jointer "
        
        if "track" in desc_lower or "channel" in desc_lower or "gs-mt" in desc_lower:
            return f"{type_str}Connector / Jointer for Magnetic Track Channel."
        return f"{type_str}Connector / Jointer."
        
    # Check for Suspension wire
    if "suspension" in desc_lower or "wire" in desc_lower:
        # Check length
        length_str = "2Mtr"
        len_match = re.search(r'(\d+)\s*(?:MTR|M|METERS)', desc_clean, re.IGNORECASE)
        if len_match:
            length_str = f"{len_match.group(1)}Mtr"
        return f"Suspension wire kit, length {length_str} with ceiling canopy and mounting clips."
        
    # Check for Power Supply / LED Driver
    if "power supply" in desc_lower or "driver" in desc_lower or "supply" in desc_lower:
        # Extract voltage
        voltage_str = "24V" # Default
        volt_match = re.search(r'(\d+V)', desc_clean, re.IGNORECASE)
        if volt_match:
            voltage_str = volt_match.group(1).upper()
            
        # Extract wattage
        wattage_str = ""
        watt_match = re.search(r'(\d+W)', desc_clean, re.IGNORECASE)
        if watt_match:
            wattage_str = watt_match.group(1).upper()
            
        volt_detail = f" {voltage_str}," if voltage_str else ""
        watt_detail = f" {wattage_str}" if wattage_str else ""
        
        # Check if IP67/waterproof
        ip_str = ""
        if "ip67" in desc_lower or "waterproof" in desc_lower or "outdoor" in desc_lower:
            ip_str = ", IP67"
            
        return f"Constant Voltage Power Supply,{volt_detail}{watt_detail}{ip_str}."
        
    # Check for Track Rail / Channel
    if "channel" in desc_lower or "track patti" in desc_lower or "track rail" in desc_lower or "track" in desc_lower:
        # Extract width
        width_str = "26"
        width_match = re.search(r'(\d+)\s*mm', desc_lower)
        if width_match:
            width_str = width_match.group(1)
            
        # Extract length
        length_str = "1000" # Default
        len_match = re.search(r'(\d+)\s*mm\s*length', desc_lower)
        if len_match:
            length_str = len_match.group(1)
        else:
            len_match2 = re.search(r'length\s*[-:\s]*\s*(\d+)', desc_lower)
            if len_match2:
                length_str = len_match2.group(1)
                
        # Check mounting type (suspended, recessed, surface)
        mount_str = "Suspended / Surface"
        if "recessed" in desc_lower or "concealed" in desc_lower:
            mount_str = "Recessed"
        elif "suspended" in desc_lower or "sus" in desc_lower:
            mount_str = "Suspended"
            
        return f"Housing - Extruded Aluminium Profile Magnetic Track Channel, Black Finish, width {width_str}mm, length {length_str}mm, {mount_str} Mounting."

    # 2. Handle standard luminaires
    # Extract Housing/Material
    housing = "Die-Cast Aluminum" # Default
    if "extruded" in desc_lower or "extruded aluminum" in desc_lower:
        housing = "Extruded Aluminum"
    elif "pc" in desc_lower or "polycarbonate" in desc_lower:
        housing = "Polycarbonate"
    elif "crca" in desc_lower:
        housing = "CRCA Powder Coated Body"
        
    # Extract Optics/Diffuser
    optics = "Reflector and Clear Glass" # Default
    if "opal" in desc_lower or "diffuser" in desc_lower or "bwf" in desc_lower:
        optics = "Opal Diffuser"
    elif "special optics" in desc_lower or "lens" in desc_lower:
        optics = "Special Optics"
        
    # Extract IP Rating
    ip = "IP20" # Default
    ip_match = re.search(r'(IP\d{2})', desc_clean, re.IGNORECASE)
    if ip_match:
        ip = ip_match.group(1).upper()
    elif "ig-" in desc_lower or "inground" in desc_lower or "insert" in desc_lower:
        ip = "IP67"
        
    # Extract Beam Angle
    beam = ""
    beam_match = re.search(r'beam\s*angle\s*[-:\s]*\s*(\d+)', desc_lower)
    if beam_match:
        beam = f"Beam Angle - {beam_match.group(1)}°"
    elif "1degree" in desc_lower or "1 degree" in desc_lower:
        beam = "Beam Angle - 1°"
    elif "30 degree" in desc_lower or "30degree" in desc_lower:
        beam = "Beam Angle - 30°"
        
    # Extract Dimensions (Dia, Height, Cutout)
    dimensions = ""
    # Check for D - XX X H - YY
    d_h_match = re.search(r'd\s*[-:\s]*\s*(\d+)\s*mm\s*[xX]\s*h\s*[-:\s]*\s*(\d+)', desc_lower)
    if d_h_match:
        dimensions = f"Dimensions - Dia - {d_h_match.group(1)}mm, Height - {d_h_match.group(2)}mm"
    else:
        # Check cut out size
        cutout_match = re.search(r'cut\s*out\s*size\s*[-:\s]*\s*(\d+)', desc_lower)
        height_match = re.search(r'height\s*[-:\s]*\s*(\d+)', desc_lower)
        if cutout_match and height_match:
            dimensions = f"Dimensions - Cutout - {cutout_match.group(1)}mm, Height - {height_match.group(1)}mm"
        elif cutout_match:
            dimensions = f"Dimensions - Cutout - {cutout_match.group(1)}mm"
            
    if not dimensions:
        # Try generic size match: size - A: 63 x B: 150
        size_match = re.search(r'size\s*[-:\s]*\s*a:\s*(\d+)\s*x\s*b:\s*(\d+)', desc_lower)
        if size_match:
            dimensions = f"Dimensions - {size_match.group(1)} x {size_match.group(2)}mm"
        else:
            # Check dia match
            dia_match = re.search(r'dia\s*[-:\s]*\s*(\d+)\s*[x*]\s*(\d+)', desc_lower)
            if dia_match:
                dimensions = f"Dimensions - Dia - {dia_match.group(1)} x {dia_match.group(2)}mm"
            else:
                dia_match_2 = re.search(r'dia\s*[-:\s]*\s*(\d+)\s*mm', desc_lower)
                if dia_match_2:
                    dimensions = f"Dimensions - Dia - {dia_match_2.group(1)}mm"
                    
    # Construct premium description
    parts = []
    if housing:
        parts.append(f"Housing - {housing}")
    if optics:
        parts.append(optics)
    if dimensions:
        parts.append(dimensions)
    if beam:
        parts.append(beam)
    if ip:
        parts.append(ip)
        
    return ", ".join(parts) + "."




@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload-boq", methods=["POST"])
def upload_boq():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
        
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "Empty filename"}), 400
        
    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active
        
        # Detect header row
        header_row_idx = None
        desc_col_idx = None
        qty_col_idx = None
        
        is_iwo = False
        iwo_cols = {}
        
        # Scan first 15 rows to find headers
        for r_idx in range(1, 16):
            row_vals = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, sheet.max_column + 1)]
            row_str = " ".join([str(v).lower() for v in row_vals if v is not None])
            
            # Check for IWO headers
            if "gls code" in row_str or "gls-code" in row_str:
                is_iwo = True
                header_row_idx = r_idx
                for idx, val in enumerate(row_vals):
                    if not val:
                        continue
                    v_clean = str(val).lower().strip()
                    if "boq description" in v_clean or "boq_description" in v_clean or "item" in v_clean:
                        iwo_cols["boq_desc"] = idx + 1
                    elif "gls code" in v_clean or "gls_code" in v_clean:
                        iwo_cols["gls_code"] = idx + 1
                    elif "specification" in v_clean or "product description" in v_clean:
                        iwo_cols["spec"] = idx + 1
                    elif "color" in v_clean:
                        iwo_cols["color"] = idx + 1
                    elif "qty" in v_clean or "quantity" in v_clean:
                        iwo_cols["qty"] = idx + 1
                    elif "unit" in v_clean:
                        iwo_cols["unit"] = idx + 1
                    elif "driver details" in v_clean or "driver_details" in v_clean:
                        iwo_cols["driver"] = idx + 1
                    elif "driver qty" in v_clean or "driver_qty" in v_clean:
                        iwo_cols["driver_qty"] = idx + 1
                    elif "led details" in v_clean or "led_details" in v_clean:
                        iwo_cols["led"] = idx + 1
                    elif "accessories" in v_clean:
                        iwo_cols["accessories"] = idx + 1
                break
                
            elif "qty" in row_str or "quantity" in row_str:
                d_col, q_col = detect_columns(row_vals)
                if d_col and q_col:
                    header_row_idx = r_idx
                    desc_col_idx = d_col
                    qty_col_idx = q_col
                    break
                
        if not header_row_idx:
            header_row_idx = 4
            desc_col_idx = 3 # Col C
            qty_col_idx = 4  # Col D
            
        items = []
        sr_no = 1
        
        if is_iwo:
            for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
                gls_code = sheet.cell(row=r_idx, column=iwo_cols.get("gls_code", 3)).value
                if not gls_code:
                    continue
                gls_code_str = str(gls_code).strip()
                if gls_code_str == "" or gls_code_str.lower().startswith("sales team") or gls_code_str.lower().startswith("total"):
                    continue
                
                boq_desc = sheet.cell(row=r_idx, column=iwo_cols.get("boq_desc", 2)).value or ""
                spec = sheet.cell(row=r_idx, column=iwo_cols.get("spec", 4)).value or ""
                color = sheet.cell(row=r_idx, column=iwo_cols.get("color", 5)).value or "Black"
                qty_val = sheet.cell(row=r_idx, column=iwo_cols.get("qty", 6)).value or 1
                unit_val = sheet.cell(row=r_idx, column=iwo_cols.get("unit", 7)).value or "Nos"
                driver = sheet.cell(row=r_idx, column=iwo_cols.get("driver", 9)).value or "Fulham"
                driver_qty_val = sheet.cell(row=r_idx, column=iwo_cols.get("driver_qty", 10)).value or 1
                led = sheet.cell(row=r_idx, column=iwo_cols.get("led", 11)).value or "Bridgelux"
                accessories = sheet.cell(row=r_idx, column=iwo_cols.get("accessories", 12)).value or "Standard"
                
                try:
                    qty = int(float(str(qty_val).strip()))
                except Exception:
                    qty = 1
                
                try:
                    driver_qty = int(float(str(driver_qty_val).strip()))
                except Exception:
                    driver_qty = 1
                    
                items.append({
                    "id": sr_no,
                    "boq_description": str(boq_desc).strip(),
                    "boq_qty": qty,
                    "unit": str(unit_val).strip(),
                    "gls_code": gls_code_str,
                    "product_description": str(spec).strip(),
                    "body_color": str(color).strip(),
                    "driver_details": str(driver).strip(),
                    "driver_qty": driver_qty,
                    "led_details": str(led).strip(),
                    "accessories": str(accessories).strip(),
                    "rate": 0,
                    "page": 0,
                    "matched_by": "iwo_upload_import"
                })
                sr_no += 1
        else:
            # Read data rows (normal BOQ mode)
            for r_idx in range(header_row_idx + 1, sheet.max_row + 1):
                desc_val = sheet.cell(row=r_idx, column=desc_col_idx).value
                qty_val = sheet.cell(row=r_idx, column=qty_col_idx).value
                
                if desc_val is not None:
                    desc_str = str(desc_val).strip()
                    if desc_str == "" or desc_str.lower().startswith("sales team") or desc_str.lower().startswith("total"):
                        continue
                    
                    # Parse Qty and Unit
                    qty = 0
                    unit = "Nos"
                    if qty_val is not None:
                        qty_str = str(qty_val).strip()
                        unit_match = re.search(r'(\d+)\s*[- ]*\s*([a-zA-Z]+)', qty_str)
                        if unit_match:
                            qty = int(unit_match.group(1))
                            unit = unit_match.group(2).capitalize()
                            if unit == "Nos" or unit == "No":
                                unit = "Nos"
                        else:
                            try:
                                qty = int(float(qty_str))
                            except ValueError:
                                qty = qty_str
                                
                    if "mtr" in desc_str.lower():
                        unit = "Mtr"
                        
                    # Run matching engine
                    parsed_info = matching_engine.parse_product_code(desc_str)
                    specs = matching_engine.lookup_catalog_database(parsed_info, CATALOG_JSON_PATH)
                    
                    product_description = specs["product_description"]
                    if product_description == "GLS-SPA Luminaire, IP20.":
                        product_description = generate_description_from_boq(desc_str)
                    
                    # Determine default body color
                    body_color = "Black"
                    if "white" in desc_str.lower() or "white" in product_description.lower():
                        body_color = "White"
                    elif "mil finish" in desc_str.lower() or "mil finish" in product_description.lower() or "silver" in desc_str.lower():
                        body_color = "Mil Finish"
                        
                    # Calculate driver details
                    driver_make = specs["driver_make"]
                    driver_wattage = specs["driver_wattage"]
                    driver_qty = qty
                    
                    if unit.lower() == "mtr":
                        driver_make = "Constant Voltage 24V"
                        driver_wattage = "150W"
                        try:
                            meters = float(qty)
                            driver_qty = max(1, int(meters / 12))
                        except Exception:
                            driver_qty = 1
                    
                    driver_details = f"{driver_make} - {driver_wattage}" if driver_make != "NA" else "NA"
                    led_details = f"{specs['led_make']} - {parsed_info.get('cct', '4000K')}" if specs["led_make"] != "NA" else "NA"
                    
                    items.append({
                        "id": sr_no,
                        "boq_description": desc_str,
                        "boq_qty": qty,
                        "unit": unit,
                        "gls_code": clean_gls_code(desc_str, specs["gls_code"]),
                        "product_description": product_description,
                        "body_color": body_color,
                        "driver_details": driver_details,
                        "driver_qty": driver_qty,
                        "led_details": led_details,
                        "accessories": specs["accessories"],
                        "page": specs["page"],
                        "matched_by": specs["matched_by"],
                        "remarks": ""
                    })
                    sr_no += 1
                
        return jsonify({
            "filename": file.filename,
            "project_name": file.filename.split("__")[0].replace("BOQ", "").replace("_", " ").strip(),
            "items": items
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to parse BOQ Excel: {str(e)}"}), 500

@app.route("/api/search-catalog", methods=["GET"])
def search_catalog():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify([])
        
    results = []
    if os.path.exists(CATALOG_JSON_PATH):
        try:
            with open(CATALOG_JSON_PATH, "r", encoding="utf-8") as f:
                catalog = json.load(f)
            for page_num, text in catalog.items():
                if query.lower() in text.lower():
                    idx = text.lower().find(query.lower())
                    start = max(0, idx - 80)
                    end = min(len(text), idx + 80)
                    snippet = text[start:end].replace("\n", " ").strip()
                    results.append({
                        "page": int(page_num),
                        "snippet": f"... {snippet} ..."
                    })
        except Exception as e:
            print("Error searching catalog:", e)
            
    results = sorted(results, key=lambda x: x["page"])
    return jsonify(results[:15])

@app.route("/api/generate-iwo", methods=["POST"])
def generate_iwo():
    data = request.json
    if not data:
        return jsonify({"error": "No data provided"}), 400
        
    client_name = data.get("client_name", "Food Square")
    iwo_no = data.get("iwo_no", "P176")
    iwo_date = data.get("iwo_date", datetime.date.today().strftime("%d/%m/%y"))
    delivery_date = data.get("delivery_date", "")
    sales_name = data.get("sales_name", "Sales-01")
    items = data.get("items", [])
    
    try:
        # Create styled Excel sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "IWO "
        ws.views.sheetView[0].showGridLines = True
        
        # Styles
        font_title_company = Font(name="Arial", size=16, bold=True)
        font_title_iwo = Font(name="Arial", size=20, bold=True)
        font_header = Font(name="Arial", size=9, bold=True)
        font_bold_data = Font(name="Arial", size=9, bold=True)
        font_data = Font(name="Arial", size=9)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_side = Side(border_style="thin", color="000000")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        col_widths = {
            "A": 3,   # Buffer column
            "B": 8,   # SR. NO.
            "C": 12,  # Image placeholder (kept empty)
            "D": 25,  # GLS product code
            "E": 45,  # product description
            "F": 12,  # Body Color
            "G": 10,  # Qty
            "H": 10,  # Unit
            "I": 25,  # DRIVER MAKE and Wattage
            "J": 12,  # DRIVER QTY
            "K": 20,  # LED MAKE and CCT
            "L": 20,  # Accessories
            "M": 22,  # stickering & branding details
            "N": 18,  # packing details
            "O": 25   # REMARKS
        }
        
        for col_let, width in col_widths.items():
            ws.column_dimensions[col_let].width = width
            
        ws.cell(row=1, column=2, value="GLS-SPA").font = font_title_company
        ws.cell(row=2, column=2, value="INTERNAL WORK ORDER").font = font_title_iwo
        
        ws.cell(row=3, column=3, value="Client Name :- ").font = font_bold_data
        ws.cell(row=3, column=4, value=client_name).font = font_bold_data
        ws.cell(row=3, column=5, value=f"IWO NO :- {iwo_no}").font = font_bold_data
        ws.cell(row=3, column=6, value="IWO DATE :-     ").font = font_bold_data
        ws.cell(row=3, column=7, value=iwo_date).font = font_bold_data
        ws.cell(row=3, column=9, value="PO NO. :-").font = font_bold_data
        ws.cell(row=3, column=12, value="DELIVERY DATE :-  ").font = font_bold_data
        ws.cell(row=3, column=13, value=delivery_date).font = font_bold_data
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[4].height = 25
        
        headers = [
            "SR. NO.", "Image", "GLS product code", "product description",
            "Body Color", "Qty", "Unit", "DRIVER MAKE and Wattage ", "DRIVER QTY",
            "LED MAKE and CCT ", "Accessories", "stickering  and branding details ",
            "packing details ", "REMARKS"
        ]
        
        for c_idx, h_text in enumerate(headers):
            cell = ws.cell(row=4, column=c_idx + 2, value=h_text)
            cell.font = font_header
            cell.border = border_all
            cell.alignment = align_center
            
        # Write data rows
        curr_row = 5
        for item in items:
            ws.row_dimensions[curr_row].height = 35 # normal, non-congested height
            
            vals = [
                item.get("id"),
                "", # Image column remains empty as requested
                item.get("gls_code"),
                item.get("product_description"),
                item.get("body_color"),
                item.get("boq_qty"),
                item.get("unit"),
                item.get("driver_details"),
                item.get("driver_qty"),
                item.get("led_details"),
                item.get("accessories") or "",
                "GLS SPA", # Default stickering
                "", # packing details remains empty
                item.get("remarks") or ""
            ]
            
            for c_idx, val in enumerate(vals):
                cell = ws.cell(row=curr_row, column=c_idx + 2, value=val)
                cell.font = font_data
                cell.border = border_all
                if c_idx in [0, 1, 4, 5, 6, 8, 9, 10, 11, 12]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
            curr_row += 1
            
        # Footer block (Signatures)
        ws.row_dimensions[curr_row + 1].height = 25
        cell_sales = ws.cell(row=curr_row + 1, column=2, value=f"SALES TEAM NAME AND SIGN  - {sales_name}")
        cell_sales.font = font_bold_data
        
        cell_prod = ws.cell(row=curr_row + 1, column=8, value="PRODUCTION SIGN.:-")
        cell_prod.font = font_bold_data
        
        cell_qc = ws.cell(row=curr_row + 1, column=12, value="QC SIGN.:-")
        cell_qc.font = font_bold_data
        
        # Save file
        filename = f"IWO-{iwo_no}-{client_name.replace(' ', '_')}.xlsx"
        save_path = os.path.join(GENERATED_FOLDER, filename)
        wb.save(save_path)
        
        return jsonify({
            "success": True,
            "filename": filename,
            "download_url": f"/api/download-iwo/{filename}"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate IWO: {str(e)}"}), 500

@app.route("/api/download-iwo/<filename>", methods=["GET"])
def download_iwo(filename):
    return send_from_directory(GENERATED_FOLDER, filename, as_attachment=True)

import csv

@app.route("/api/database", methods=["GET"])
def get_database():
    csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GLS_SPA_Product_Database.csv")
    if not os.path.exists(csv_path):
        return jsonify({"error": "Database not found"}), 404
    try:
        products = []
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                products.append(row)
        return jsonify(products)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/database", methods=["POST"])
def save_database():
    data = request.json
    if not isinstance(data, list):
        return jsonify({"error": "Invalid data format"}), 400
        
    csv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GLS_SPA_Product_Database.csv")
    catalog_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GLS_SPA_catalog.json")
    html_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "GLS_SPA_Product_Database.html")
    html_copy_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "catlogues", "GLS_SPA_Product_Database.html")
    
    try:
        # 1. Save to CSV
        headers = [
            "Product Name", "Product Code", "Size", "Lens", 
            "Driver", "Beam Angle", "IP Rating", "Housing", 
            "Driver Make", "Driver Wattage"
        ]
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for p in data:
                writer.writerow([
                    p.get("Product Name", p.get("product_name", p.get("name", ""))),
                    p.get("Product Code", p.get("code", "")),
                    p.get("Size", p.get("size", "")),
                    p.get("Lens", p.get("lens", "")),
                    p.get("Driver", p.get("driver", "Yes")),
                    p.get("Beam Angle", p.get("beam_angle", "")),
                    p.get("IP Rating", p.get("ip_rating", "")),
                    p.get("Housing", p.get("housing", "")),
                    p.get("Driver Make", p.get("driver_make", "")),
                    p.get("Driver Wattage", p.get("driver_wattage", ""))
                ])
                
        # 2. Sync to GLS_SPA_catalog.json
        existing_images = {}
        if os.path.exists(catalog_path):
            with open(catalog_path, "r", encoding="utf-8") as f:
                cat_data = json.load(f)
                for item in cat_data:
                    code = item.get("code", "").strip().upper()
                    if code and item.get("b64"):
                        existing_images[code] = item.get("b64")
                        
        catalog_products = []
        for p in data:
            code = p.get("Product Code", p.get("code", "")).strip()
            if not code:
                continue
            norm_code = code.upper()
            b64 = existing_images.get(norm_code, "")
            
            # Retrieve page or fallback
            page_val = p.get("page", p.get("pdf_page", 0))
            try:
                page = int(page_val)
            except Exception:
                page = 0
                
            catalog_products.append({
                "product_code": code,
                "code": code,
                "pdf_page": page,
                "page": page,
                "catalog_name": p.get("catalog_name", "INDOOR SERIES 2026.pdf"),
                "filename": p.get("filename", f"page_{page}.jpeg"),
                "size": p.get("Size", p.get("size", "1000x1000")),
                "b64": b64
            })
            
        with open(catalog_path, "w", encoding="utf-8") as f:
            json.dump(catalog_products, f, indent=2, ensure_ascii=False)
            
        # 3. Rebuild DATA array in HTML files
        def update_html(p_path):
            if not os.path.exists(p_path):
                return
            with open(p_path, "r", encoding="utf-8") as h_f:
                html = h_f.read()
            pattern = r'const\s+DATA\s*=\s*\[.*?\];'
            js_array = json.dumps(catalog_products, ensure_ascii=False)
            new_html = re.sub(pattern, f"const DATA = {js_array};", html, flags=re.DOTALL)
            with open(p_path, "w", encoding="utf-8") as h_f:
                h_f.write(new_html)
                
        update_html(html_path)
        update_html(html_copy_path)
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def num_to_words_indian(num):
    if num == 0:
        return "Zero Only"
    rupees = int(num)
    paise = int(round((num - rupees) * 100))
    
    under_20 = ['Zero', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine', 'Ten',
                'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
    tens = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
    
    def helper(n):
        if n == 0: return ""
        if n < 20:
            return under_20[int(n)]
        elif n < 100:
            tens_digit = int(n / 10)
            ones_digit = int(n % 10)
            return tens[tens_digit] + (' ' + under_20[ones_digit] if ones_digit else '')
        elif n < 1000:
            hundreds_digit = int(n / 100)
            rem = n % 100
            return under_20[hundreds_digit] + ' Hundred' + (' ' + helper(rem) if rem else '')
            
    def get_parts(n):
        if n == 0: return "Zero"
        crore = int(n / 10000000)
        n %= 10000000
        lakh = int(n / 100000)
        n %= 100000
        thousand = int(n / 1000)
        n %= 1000
        hundred = int(n)
        
        parts = []
        if crore:
            parts.append(helper(crore) + " Crore")
        if lakh:
            parts.append(helper(lakh) + " Lakh")
        if thousand:
            parts.append(helper(thousand) + " Thousand")
        if hundred:
            parts.append(helper(hundred))
        return " ".join(parts).strip()
        
    rupees_text = get_parts(rupees) + " Rupees" if rupees > 0 else "Zero Rupees"
    if paise > 0:
        paise_text = get_parts(paise) + " Paise"
        return rupees_text + " and " + paise_text + " Only"
    return rupees_text + " Only"

@app.route("/api/generate-invoice", methods=["POST"])
def generate_invoice():
    try:
        req_data = request.json
        items = req_data.get("items", [])
        
        # Buyer details
        buyer_name = req_data.get("buyer_name", "Shahi lites & Chandeliers").strip()
        buyer_address = req_data.get("buyer_address", "Faizabad Road, Lucknow").strip()
        buyer_gstin = req_data.get("buyer_gstin", "09AEQPM4658E1Z1").strip()
        buyer_contact = req_data.get("buyer_contact", "ZUBERUL HUSSAIN").strip()
        
        # Invoice metadata
        invoice_no = req_data.get("invoice_no", "SOR/26-27/001").strip()
        invoice_date = req_data.get("invoice_date", "29/06/26").strip()
        payment_terms = req_data.get("payment_terms", "100% Advance").strip()
        validity = req_data.get("validity", "7 days").strip()
        destination = req_data.get("destination", "Lucknow").strip()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proforma Invoice"
        ws.views.sheetView[0].showGridLines = True
        
        # Setup column widths
        widths = {
            'A': 6,   # Sr No
            'B': 38,  # Description
            'C': 12,  # HSN
            'D': 8,   # Qty
            'E': 12,  # Price
            'F': 12,  # Amount
            'G': 8,   # GST (%)
            'H': 12,  # GST Amount
            'I': 14   # Grand Total
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
            
        # Fonts and Styles
        font_brand = Font(name="Playfair Display", size=18, bold=True, color="8BC34A")
        font_brand_address = Font(name="Inter", size=9, color="535b6d")
        font_title = Font(name="Playfair Display", size=14, bold=True, italic=True, color="FFFFFF")
        font_bold = Font(name="Inter", size=10, bold=True)
        font_regular = Font(name="Inter", size=9)
        
        fill_brand_header = PatternFill(start_color="151924", end_color="151924", fill_type="solid")
        fill_table_header = PatternFill(start_color="070812", end_color="070812", fill_type="solid")
        font_table_header = Font(name="Inter", size=9, bold=True, color="8BC34A")
        
        border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Row 1-3: GLS-SPA Branding Header
        ws.merge_cells("A1:I2")
        cell_brand = ws["A1"]
        cell_brand.value = "GLS-SPA LIGHTING"
        cell_brand.font = font_brand
        cell_brand.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells("A3:I3")
        cell_brand_add = ws["A3"]
        cell_brand_add.value = "Gala No. 202, Bliss Nirman Industrial Estate, Gokhiware, Vasai East, Palghar, Maharashtra - 401208 | enquiry@svldpl.net"
        cell_brand_add.font = font_brand_address
        cell_brand_add.alignment = Alignment(horizontal="center", vertical="center")
        
        # Row 5: Proforma Invoice Title
        ws.merge_cells("A5:I5")
        cell_title = ws["A5"]
        cell_title.value = "PROFORMA INVOICE"
        cell_title.font = font_title
        cell_title.fill = fill_brand_header
        cell_title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 25
        
        # Row 7-12: Buyer & Invoice Metadata Blocks
        # Left (Buyer)
        ws["A7"] = "Buyer (Bill to)"
        ws["A7"].font = font_bold
        ws["A8"] = buyer_name
        ws["A8"].font = font_bold
        ws["A9"] = f"Address: {buyer_address}"
        ws["A9"].font = font_regular
        ws["A10"] = f"GSTIN/UIN: {buyer_gstin}"
        ws["A10"].font = font_regular
        ws["A11"] = f"Contact: {buyer_contact}"
        ws["A11"].font = font_regular
        
        # Right (Invoice Info)
        ws["F7"] = "Voucher No."
        ws["F7"].font = font_bold
        ws["G7"] = invoice_no
        ws["G7"].font = font_regular
        
        ws["F8"] = "Date"
        ws["F8"].font = font_bold
        ws["G8"] = invoice_date
        ws["G8"].font = font_regular
        
        ws["F9"] = "Payment Terms"
        ws["F9"].font = font_bold
        ws["G9"] = payment_terms
        ws["G9"].font = font_regular
        
        ws["F10"] = "Validity Period"
        ws["F10"].font = font_bold
        ws["G10"] = validity
        ws["G10"].font = font_regular
        
        ws["F11"] = "Destination"
        ws["F11"].font = font_bold
        ws["G11"] = destination
        ws["G11"].font = font_regular
        
        # Row 13: Table Headers
        headers = [
            "Sr No", "Description", "HSN Code", "Qty", "Price", 
            "Amount", "GST (%)", "GST Amount", "Grand Total"
        ]
        ws.row_dimensions[13].height = 24
        for idx, h in enumerate(headers):
            col_letter = get_column_letter(idx + 1)
            cell = ws[f"{col_letter}13"]
            cell.value = h
            cell.font = font_table_header
            cell.fill = fill_table_header
            cell.alignment = Alignment(horizontal="center" if idx != 1 else "left", vertical="center")
            cell.border = border_thin
            
        # Data rows
        curr_row = 14
        total_qty = 0
        total_amount = 0
        total_gst_amt = 0
        total_grand = 0
        
        for idx, item in enumerate(items):
            ws.row_dimensions[curr_row].height = 20
            
            rate_val = 0
            try:
                rate_val = float(item.get("rate", item.get("price", 0)))
            except Exception:
                pass
                
            qty_val = 0
            try:
                qty_val = int(item.get("boq_qty", 0))
            except Exception:
                pass
                
            amount_val = qty_val * rate_val
            gst_val = amount_val * 0.18
            grand_val = amount_val + gst_val
            
            total_qty += qty_val
            total_amount += amount_val
            total_gst_amt += gst_val
            total_grand += grand_val
            
            ws[f"A{curr_row}"] = idx + 1
            ws[f"B{curr_row}"] = item.get("gls_code", "")
            ws[f"C{curr_row}"] = "9405"
            ws[f"D{curr_row}"] = qty_val
            ws[f"E{curr_row}"] = rate_val
            ws[f"F{curr_row}"] = amount_val
            ws[f"G{curr_row}"] = "18%"
            ws[f"H{curr_row}"] = gst_val
            ws[f"I{curr_row}"] = grand_val
            
            # Formats and alignments
            for c_idx in range(1, 10):
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.font = font_regular
                cell.border = border_thin
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
            curr_row += 1
            
        # Totals Row
        ws.row_dimensions[curr_row].height = 22
        ws[f"A{curr_row}"] = "Total"
        ws[f"A{curr_row}"].font = font_bold
        ws.merge_cells(f"A{curr_row}:C{curr_row}")
        
        ws[f"D{curr_row}"] = total_qty
        ws[f"D{curr_row}"].font = font_bold
        ws[f"F{curr_row}"] = total_amount
        ws[f"F{curr_row}"].font = font_bold
        ws[f"H{curr_row}"] = total_gst_amt
        ws[f"H{curr_row}"].font = font_bold
        ws[f"I{curr_row}"] = total_grand
        ws[f"I{curr_row}"].font = font_bold
        
        for c_idx in range(1, 10):
            cell = ws.cell(row=curr_row, column=c_idx)
            cell.border = border_thin
            if c_idx not in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        # Word Amount Row
        curr_row += 2
        ws.merge_cells(f"A{curr_row}:I{curr_row}")
        ws[f"A{curr_row}"] = f"Amount Chargeable (in words): {num_to_words_indian(int(total_grand))}"
        ws[f"A{curr_row}"].font = font_bold
        
        # Bank details
        curr_row += 2
        ws.merge_cells(f"A{curr_row}:D{curr_row}")
        ws[f"A{curr_row}"] = "Company's Bank Details"
        ws[f"A{curr_row}"].font = font_bold
        
        ws[f"A{curr_row+1}"] = "Bank Name: HDFC Bank"
        ws[f"A{curr_row+1}"].font = font_regular
        ws[f"A{curr_row+2}"] = "A/c No.: 50200034798292"
        ws[f"A{curr_row+2}"].font = font_regular
        ws[f"A{curr_row+3}"] = "Branch & IFSC: Vasai East & HDFC0000038"
        ws[f"A{curr_row+3}"].font = font_regular
        
        # Terms Block
        curr_row += 5
        ws[f"A{curr_row}"] = "Terms and Conditions:"
        ws[f"A{curr_row}"].font = font_bold
        
        terms = [
            "1. 2 years warranty from date of dispatch against manufacturing defects.",
            "2. Cancellation: Orders cannot be cancelled once confirmed.",
            "3. Title of goods: Transfers to buyer only upon full payment.",
            "4. No returns without prior written approval.",
            "5. These custom orders are non-returnable."
        ]
        for t_idx, t in enumerate(terms):
            ws[f"A{curr_row + t_idx + 1}"] = t
            ws[f"A{curr_row + t_idx + 1}"].font = font_regular
            
        # Signature block
        ws[f"G{curr_row}"] = "For GLS-SPA LIGHTING"
        ws[f"G{curr_row}"].font = font_bold
        ws[f"G{curr_row+4}"] = "Authorised Signatory"
        ws[f"G{curr_row+4}"].font = font_bold
        
        filename = f"PI-{invoice_no.replace('/', '_')}-{buyer_name.replace(' ', '_')}.xlsx"
        save_path = os.path.join(GENERATED_FOLDER, filename)
        wb.save(save_path)
        
        return jsonify({
            "success": True,
            "filename": filename,
            "download_url": f"/api/download-iwo/{filename}"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate Invoice: {str(e)}"}), 500


@app.route("/api/buyers", methods=["GET", "POST"])
def manage_buyers():
    if request.method == "POST":
        try:
            new_buyer = request.json
            if not new_buyer.get("name"):
                return jsonify({"error": "Buyer name is required"}), 400
                
            buyers = []
            if os.path.exists(BUYERS_DB_PATH):
                with open(BUYERS_DB_PATH, "r", encoding="utf-8") as f:
                    try:
                        buyers = json.load(f)
                    except Exception:
                        buyers = []
            
            # Check if exists, update or append
            exists = False
            for idx, b in enumerate(buyers):
                if b.get("name").strip().lower() == new_buyer.get("name").strip().lower():
                    buyers[idx] = new_buyer
                    exists = True
                    break
            if not exists:
                buyers.append(new_buyer)
                
            with open(BUYERS_DB_PATH, "w", encoding="utf-8") as f:
                json.dump(buyers, f, indent=2)
                
            return jsonify({"success": True, "buyers": buyers})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    else:
        buyers = []
        if os.path.exists(BUYERS_DB_PATH):
            with open(BUYERS_DB_PATH, "r", encoding="utf-8") as f:
                try:
                    buyers = json.load(f)
                except Exception:
                    buyers = []
        return jsonify(buyers)

@app.route("/api/generate-invoice-pdf", methods=["POST"])
def generate_invoice_pdf():
    try:
        req_data = request.json
        items = req_data.get("items", [])
        
        # Buyer details
        buyer_name = req_data.get("buyer_name", "").strip()
        buyer_address = req_data.get("buyer_address", "").strip()
        buyer_gstin = req_data.get("buyer_gstin", "").strip()
        buyer_contact = req_data.get("buyer_contact", "").strip()
        
        # Invoice metadata
        invoice_no = req_data.get("invoice_no", "SOR/26-27/001").strip()
        invoice_date = req_data.get("invoice_date", "").strip()
        payment_terms = req_data.get("payment_terms", "100% Advance").strip()
        validity = req_data.get("validity", "7 days").strip()
        destination = req_data.get("destination", "").strip()
        
        # Build PDF path
        filename = f"PI-{invoice_no.replace('/', '_')}-{buyer_name.replace(' ', '_')}.pdf"
        pdf_path = os.path.join(GENERATED_FOLDER, filename)
        
        # Setup document
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            leftMargin=30,
            rightMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        style_title = ParagraphStyle(
            'InvoiceTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            textColor=colors.HexColor('#8BC34A'), # Apple green
            alignment=1, # Center
            spaceAfter=5
        )
        style_subtitle = ParagraphStyle(
            'InvoiceSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#535b6d'),
            alignment=1, # Center
            spaceAfter=15
        )
        style_section_title = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=12,
            textColor=colors.white,
            backColor=colors.HexColor('#151924'),
            borderPadding=6,
            spaceAfter=12,
            alignment=1
        )
        style_label = ParagraphStyle(
            'Label',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.HexColor('#151924')
        )
        style_val = ParagraphStyle(
            'Value',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor('#333333')
        )
        style_table_header = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=colors.HexColor('#8BC34A'),
            alignment=1
        )
        style_table_cell = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            alignment=1
        )
        style_table_cell_left = ParagraphStyle(
            'TableCellLeft',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            alignment=0
        )
        style_table_cell_bold = ParagraphStyle(
            'TableCellBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=8,
            alignment=1
        )
        
        story = []
        
        # Company Info Header
        story.append(Paragraph("GLS-SPA LIGHTING", style_title))
        story.append(Paragraph("Gala No. 202, Bliss Nirman Industrial Estate, Building No. 5, Gokhiware, Vasai East, Palghar, Maharashtra - 401208 | enquiry@svldpl.net", style_subtitle))
        story.append(Paragraph("PROFORMA INVOICE", style_section_title))
        
        # Meta Block: Buyer vs PI details
        meta_data = [
            [
                Paragraph("<b>Buyer (Bill to)</b>", style_label), "",
                Paragraph("<b>Voucher Details</b>", style_label), ""
            ],
            [
                Paragraph("Name:", style_label), Paragraph(buyer_name, style_val),
                Paragraph("Voucher No:", style_label), Paragraph(invoice_no, style_val)
            ],
            [
                Paragraph("Address:", style_label), Paragraph(buyer_address, style_val),
                Paragraph("Date:", style_label), Paragraph(invoice_date, style_val)
            ],
            [
                Paragraph("GSTIN:", style_label), Paragraph(buyer_gstin, style_val),
                Paragraph("Payment Terms:", style_label), Paragraph(payment_terms, style_val)
            ],
            [
                Paragraph("Contact:", style_label), Paragraph(buyer_contact, style_val),
                Paragraph("Validity:", style_label), Paragraph(validity, style_val)
            ],
            [
                "", "",
                Paragraph("Destination:", style_label), Paragraph(destination, style_val)
            ]
        ]
        
        meta_table = Table(meta_data, colWidths=[65, 200, 95, 175])
        meta_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (1, 0)),
            ('SPAN', (2, 0), (3, 0)),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('LINEBELOW', (0, 0), (1, 0), 1, colors.HexColor('#CCCCCC')),
            ('LINEBELOW', (2, 0), (3, 0), 1, colors.HexColor('#CCCCCC')),
        ]))
        
        story.append(meta_table)
        story.append(Spacer(1, 15))
        
        # Items Table
        table_data = [[
            Paragraph("Sr.", style_table_header),
            Paragraph("Description", style_table_header),
            Paragraph("HSN Code", style_table_header),
            Paragraph("Qty", style_table_header),
            Paragraph("Price", style_table_header),
            Paragraph("Amount", style_table_header),
            Paragraph("GST (%)", style_table_header),
            Paragraph("GST Amt", style_table_header),
            Paragraph("Grand Total", style_table_header)
        ]]
        
        total_qty = 0
        total_amount = 0
        total_gst_amt = 0
        total_grand = 0
        
        for idx, item in enumerate(items):
            rate_val = 0
            try:
                rate_val = float(item.get("rate", item.get("price", 0)))
            except Exception:
                pass
                
            qty_val = 0
            try:
                qty_val = int(item.get("boq_qty", 0))
            except Exception:
                pass
                
            amount_val = qty_val * rate_val
            
            gst_percent = 18.0
            try:
                gst_percent = float(item.get("gst_percent", 18.0))
            except Exception:
                pass
                
            gst_val = amount_val * (gst_percent / 100)
            grand_val = amount_val + gst_val
            
            total_qty += qty_val
            total_amount += amount_val
            total_gst_amt += gst_val
            total_grand += grand_val
            
            hsn = item.get("hsn_code", "9405").strip() or "9405"
            
            table_data.append([
                Paragraph(str(idx + 1), style_table_cell),
                Paragraph(item.get("gls_code", ""), style_table_cell_left),
                Paragraph(hsn, style_table_cell),
                Paragraph(str(qty_val), style_table_cell),
                Paragraph(f"Rs. {rate_val:,.2f}", style_table_cell),
                Paragraph(f"Rs. {amount_val:,.2f}", style_table_cell),
                Paragraph(f"{int(gst_percent) if gst_percent.is_integer() else gst_percent}%", style_table_cell),
                Paragraph(f"Rs. {gst_val:,.2f}", style_table_cell),
                Paragraph(f"Rs. {grand_val:,.2f}", style_table_cell)
            ])
            
        # Totals Row
        table_data.append([
            Paragraph("<b>Total</b>", style_table_cell_bold), "", "",
            Paragraph(f"<b>{total_qty}</b>", style_table_cell_bold), "",
            Paragraph(f"<b>Rs. {total_amount:,.2f}</b>", style_table_cell_bold), "",
            Paragraph(f"<b>Rs. {total_gst_amt:,.2f}</b>", style_table_cell_bold),
            Paragraph(f"<b>Rs. {total_grand:,.2f}</b>", style_table_cell_bold)
        ])
        
        # ColWidths sum: 30 + 175 + 50 + 35 + 55 + 55 + 35 + 50 + 50 = 535
        items_table = Table(table_data, colWidths=[25, 150, 45, 30, 55, 60, 35, 60, 75])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#070812')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
            ('SPAN', (0, -1), (2, -1)), # Merge Totals header
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        story.append(items_table)
        story.append(Spacer(1, 15))
        
        # Grand Total in words & Bank & Terms KeepTogether block
        lower_block = []
        lower_block.append(Paragraph(f"<b>Amount Chargeable (in words):</b> {num_to_words_indian(total_grand)}", style_val))
        lower_block.append(Spacer(1, 15))
        
        # Bank & Signature & Terms grid
        bank_terms_data = [
            [
                Paragraph("<b>Company's Bank Details</b>", style_label), "",
                Paragraph("<b>For GLS-SPA LIGHTING</b>", style_label)
            ],
            [
                Paragraph("Bank Name: HDFC Bank", style_val), "",
                ""
            ],
            [
                Paragraph("A/c No: 50200034798292", style_val), "",
                ""
            ],
            [
                Paragraph("Branch & IFSC: Vasai East & HDFC0000038", style_val), "",
                Paragraph("Authorised Signatory", style_label)
            ]
        ]
        
        bank_terms_table = Table(bank_terms_data, colWidths=[200, 135, 200])
        bank_terms_table.setStyle(TableStyle([
            ('SPAN', (0, 0), (1, 0)),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        lower_block.append(bank_terms_table)
        lower_block.append(Spacer(1, 15))
        
        # Terms and conditions
        lower_block.append(Paragraph("<b>Terms and Conditions:</b>", style_label))
        terms_text = (
            "1. 2 years warranty from date of dispatch against manufacturing defects.<br/>"
            "2. Cancellation: Orders cannot be cancelled once confirmed.<br/>"
            "3. Title of goods: Transfers to buyer only upon full payment.<br/>"
            "4. No returns without prior written approval.<br/>"
            "5. These custom orders are non-returnable."
        )
        lower_block.append(Paragraph(terms_text, style_val))
        
        story.append(KeepTogether(lower_block))
        
        # Build document
        doc.build(story)
        
        return jsonify({
            "success": True,
            "filename": filename,
            "download_url": f"/api/download-iwo/{filename}"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate PDF Invoice: {str(e)}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=True)
